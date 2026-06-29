"""
SISTEMA EXPERTO DE DATA CENTERS v2.1 FUSIONADO
Interfaz Visual Premium + Motor Prolog v2.0 Avanzado
Características: Tarjetas visuales, Comparador 3 opciones, TCO completo, Redundancia, Tier, Workloads, Ranking
"""

import customtkinter as ctk
import ollama
from pyswip import Prolog
import json
import threading
import os
import re
from datetime import datetime
from dataclasses import dataclass, field
from typing import List, Dict, Optional, Tuple
from PIL import Image, ImageOps
import matplotlib.pyplot as plt
import traceback
from pathlib import Path

try:
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# ============================================================
# CONFIGURACIÓN GLOBAL
# ============================================================

BASE_DIR = Path(__file__).parent.absolute()
RUTA_PROLOG_FIJA = Path(r"C:\Escritorio\PRO_PROLOG\PRO_PROLOG\arquitecto_gui.pl")
RUTA_PROLOG_LOCAL = BASE_DIR / "arquitecto_gui.pl"
RUTA_REPORTES = BASE_DIR / "reportes"
RUTA_HISTORIAL = BASE_DIR / "historial"
RUTA_TEMP = BASE_DIR / "temp"

for ruta in [RUTA_REPORTES, RUTA_HISTORIAL, RUTA_TEMP]:
    ruta.mkdir(exist_ok=True)

ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("blue")

TIPO_CAMBIO_USD_MXN = 17.00
COLOR_EXITO = "#22c55e"
COLOR_ADVERTENCIA = "#f59e0b"
COLOR_ERROR = "#ef4444"
COLOR_INFO = "#00ffcc"
COLOR_DINERO = "#ffaa00"

imagenes_cache = []

# ============================================================
# ESTRUCTURAS DE DATOS
# ============================================================

@dataclass
class ContextoUsuario:
    """Contexto expandido del usuario con soporte para funciones avanzadas"""
    cores_min: int = 0
    ram_min: int = 0
    iops_min: int = 0
    gpu_vram_min: int = 0
    presupuesto: int = 0
    workload: str = ""
    tier: str = "tier2"
    redundancia: str = "ninguna"
    anios_amortizacion: int = 5
    historial_chat: List[Dict] = field(default_factory=list)
    ultima_arquitectura: Optional[Dict] = None
    
    def limpiar(self):
        self.cores_min = self.ram_min = self.iops_min = self.gpu_vram_min = self.presupuesto = 0
        self.workload = ""
        self.ultima_arquitectura = None

# ============================================================
# MOTOR PROLOG V2.0
# ============================================================

class MotorProlog:
    """Interfaz con Prolog v2.0 con todas las funcionalidades avanzadas"""
    
    def __init__(self):
        self.prolog = Prolog()
        self.cargado = False
        self._cargar_prolog()
    
    def _cargar_prolog(self):
        """Cargar base de conocimiento Prolog con respaldo de rutas"""
        try:
            if RUTA_PROLOG_FIJA.exists():
                self.prolog.consult(str(RUTA_PROLOG_FIJA))
                self.cargado = True
                print(f"✅ Prolog cargado: {RUTA_PROLOG_FIJA}")
            elif RUTA_PROLOG_LOCAL.exists():
                self.prolog.consult(str(RUTA_PROLOG_LOCAL))
                self.cargado = True
                print(f"✅ Prolog cargado: {RUTA_PROLOG_LOCAL}")
            else:
                print(f"❌ Archivo Prolog no encontrado en ninguna ubicación")
        except Exception as e:
            print(f"❌ Error cargando Prolog: {e}")
            traceback.print_exc()
    
    def arquitectura_completa(self, cores: int, ram: int, iops: int, gpu: int, pres: int,
                              redundancia: str = "ninguna", anios: int = 5) -> List[Dict]:
        """Consulta avanzada con límite de búsqueda para evitar bucles"""
        try:
            # Añadimos un límite artificial de 100 resultados para romper el bucle
            query = f"""arquitectura_completa({cores}, {ram}, {iops}, {gpu}, {pres}, 
                       {redundancia}, {anios}, CAPEX, TCO, Tier, Arq), 
                       limit(100, (true))""" 
            return list(self.prolog.query(query))
        except Exception as e:
            print(f"Error en arquitectura_completa: {e}")
            return []
    
    def consulta_ranking(self, n: int, cores: int, ram: int, iops: int, gpu: int, pres: int) -> List[Dict]:
        """Obtener Top-N arquitecturas por score"""
        try:
            query = f"consulta_ranking({n}, {cores}, {ram}, {iops}, {gpu}, {pres}, TopN)"
            return list(self.prolog.query(query))
        except Exception as e:
            print(f"Error en ranking: {e}")
            return []
    
    def consulta_workload(self, workload: str, pres: int, anios: int = 5) -> List[Dict]:
        """Diseñar para perfil de carga específico"""
        try:
            query = f"consulta_workload({workload}, {pres}, {anios}, Resultado)"
            return list(self.prolog.query(query))
        except Exception as e:
            print(f"Error en workload: {e}")
            return []
    
    def diagnostico(self, cores: int, ram: int, iops: int, gpu: int, pres: int):
        """Ejecutar diagnóstico de por qué no hay solución"""
        try:
            query = f"diagnostico({cores}, {ram}, {iops}, {gpu}, {pres})"
            list(self.prolog.query(query))
        except Exception as e:
            print(f"Error en diagnóstico: {e}")
    
    def disenar_cluster(self, cores: int, ram: int, iops: int, gpu: int, pres: int,
                       redundancia: str = "ninguna", anios: int = 5) -> List[Dict]:
        """Diseñar cluster completo con TCO"""
        try:
            query = f"""consulta_cluster({cores}, {ram}, {iops}, {gpu}, {pres}, 
                       {redundancia}, {anios}, Resultado)"""
            return list(self.prolog.query(query))
        except Exception as e:
            print(f"Error en cluster: {e}")
            return []

# ============================================================
# UTILIDADES DE PARSEO Y FORMATO
# ============================================================

def limpiar_valor(valor: str) -> str:
    """Limpiar valores de Prolog"""
    return str(valor).strip().replace("'", "").replace('"', '')

def buscar_imagen(ruta_relativa: str) -> Optional[Path]:
    """
    Buscar imagen desde rutas relativas del catálogo Prolog.
    Las rutas vienen como 'imagenes/cpu/amd_epyc_9654.jpg'
    """
    if not ruta_relativa:
        return None
    
    ruta_relativa = limpiar_valor(ruta_relativa).replace("\\", "/").strip()
    if not ruta_relativa:
        return None
    
    # Separar extensión para probar alternativas
    if "." in ruta_relativa.split("/")[-1]:
        base_sin_ext, ext_original = ruta_relativa.rsplit(".", 1)
        ext_original = "." + ext_original
    else:
        base_sin_ext = ruta_relativa
        ext_original = ""
    
    # Candidatos en orden de prioridad
    candidatos = [
        BASE_DIR / ruta_relativa,
        Path(ruta_relativa),
    ]
    
    # Extensiones alternativas
    for ext in [".jpg", ".jpeg", ".png", ".webp", ".JPG", ".JPEG", ".PNG"]:
        if ext != ext_original:
            candidatos.append(BASE_DIR / f"{base_sin_ext}{ext}")
    
    for ruta in candidatos:
        try:
            if ruta.exists():
                return ruta
        except Exception:
            pass
    
    return None

def formatear_precio(precio: float, moneda: str = "USD") -> str:
    """Formatear precio con moneda"""
    if moneda == "MXN":
        precio *= TIPO_CAMBIO_USD_MXN
        return f"${precio:,.0f} MXN"
    return f"${precio:,.0f} USD"

def parsear_arquitectura_41_campos(arq_str: str) -> Dict:
    """
    Parsear estructura arq(...) de 41 campos del Prolog v2.0
    
    Estructura:
    - Campos 0-30: Datos técnicos (IDs, specs, métricas)
    - Campos 31-40: Rutas de imágenes de los 10 componentes
    """
    try:
        contenido = arq_str.replace("arq(", "").rstrip(")")
        partes = []
        nivel_paren = 0
        actual = ""
        
        # Parser robusto que respeta paréntesis anidados
        for char in contenido:
            if char in "([{":
                nivel_paren += 1
            elif char in ")]}":
                nivel_paren -= 1
            elif char == "," and nivel_paren == 0:
                partes.append(limpiar_valor(actual))
                actual = ""
                continue
            actual += char
        if actual:
            partes.append(limpiar_valor(actual))
        
        # Validar que tenemos 41 campos
        if len(partes) < 41:
            print(f"⚠️ Advertencia: Se esperaban 41 campos, se obtuvieron {len(partes)}")
        
        # Mapeo completo de los 41 campos
        return {
            # Datos técnicos (campos 0-30)
            "cpu_id": partes[0] if len(partes) > 0 else "?",
            "cpu_brand": partes[1] if len(partes) > 1 else "?",
            "cpu_model": partes[2] if len(partes) > 2 else "?",
            "cores": int(partes[3]) if len(partes) > 3 else 0,
            "mb_id": partes[4] if len(partes) > 4 else "?",
            "ram_id": partes[5] if len(partes) > 5 else "?",
            "ram_gb": int(partes[6]) if len(partes) > 6 else 0,
            "storage_id": partes[7] if len(partes) > 7 else "?",
            "storage_type": partes[8] if len(partes) > 8 else "?",
            "storage_tb": float(partes[9]) if len(partes) > 9 and partes[9] not in ("", "?") else 0.0,
            "iops": int(partes[10]) if len(partes) > 10 else 0,
            "gpu_id": partes[11] if len(partes) > 11 else "g0",
            "gpu_model": partes[12] if len(partes) > 12 else "Sin GPU",
            "vram": int(partes[13]) if len(partes) > 13 else 0,
            "gpu_tdp": int(partes[14]) if len(partes) > 14 else 0,
            "nic_id": partes[15] if len(partes) > 15 else "?",
            "nic_model": partes[16] if len(partes) > 16 else "?",
            "nic_gbps": int(partes[17]) if len(partes) > 17 else 0,
            "latency_us": float(partes[18]) if len(partes) > 18 and partes[18] not in ("", "?") else 0.0,
            "cool_id": partes[19] if len(partes) > 19 else "?",
            "cool_type": partes[20] if len(partes) > 20 else "?",
            "pue": float(partes[21]) if len(partes) > 21 and partes[21] not in ("", "?") else 1.5,
            "psu_id": partes[22] if len(partes) > 22 else "?",
            "psu_watts": int(partes[23]) if len(partes) > 23 else 0,
            "psu_eff": float(partes[24]) if len(partes) > 24 and partes[24] not in ("", "?") else 0.9,
            "ups_id": partes[25] if len(partes) > 25 else "?",
            "ups_auto_min": int(partes[26]) if len(partes) > 26 else 0,
            "chasis_id": partes[27] if len(partes) > 27 else "?",
            "rack_u": int(partes[28]) if len(partes) > 28 else 0,
            "tdp_total": int(partes[29]) if len(partes) > 29 else 0,
            "consumo_real": float(partes[30]) if len(partes) > 30 and partes[30] not in ("", "?") else 0.0,
            
            # Rutas de imágenes (campos 31-40) - Los 10 componentes
            "img_cpu": partes[31] if len(partes) > 31 else "",
            "img_mb": partes[32] if len(partes) > 32 else "",
            "img_ram": partes[33] if len(partes) > 33 else "",
            "img_storage": partes[34] if len(partes) > 34 else "",
            "img_gpu": partes[35] if len(partes) > 35 else "",
            "img_nic": partes[36] if len(partes) > 36 else "",
            "img_cool": partes[37] if len(partes) > 37 else "",
            "img_psu": partes[38] if len(partes) > 38 else "",
            "img_ups": partes[39] if len(partes) > 39 else "",
            "img_chasis": partes[40] if len(partes) > 40 else "",
        }
    except Exception as e:
        print(f"❌ Error parseando arquitectura: {e}")
        traceback.print_exc()
        return {}

def crear_imagen_ctk(ruta: Optional[Path], tamaño: Tuple[int, int] = (210, 90)) -> Optional[ctk.CTkImage]:
    """Crear imagen CTK desde ruta con corrección EXIF"""
    try:
        if not ruta or not ruta.exists():
            return None
        
        img = Image.open(ruta)
        img = ImageOps.exif_transpose(img)  # Corregir rotación EXIF
        img = img.convert("RGBA") if img.mode in ("P", "LA") else img
        img.thumbnail(tamaño, Image.LANCZOS)
        
        # Canvas del tamaño exacto
        canvas = Image.new("RGBA", tamaño, (255, 255, 255, 255))
        offset = ((tamaño[0] - img.width) // 2, (tamaño[1] - img.height) // 2)
        canvas.paste(img, offset)
        
        img_ctk = ctk.CTkImage(light_image=canvas, dark_image=canvas, size=tamaño)
        imagenes_cache.append(img_ctk)  # Mantener referencia viva
        return img_ctk
    except Exception as e:
        print(f"⚠️ No se pudo cargar imagen {ruta}: {e}")
        return None

def preparar_imagen_pdf(ruta_img: str) -> Optional[str]:
    """Convertir imagen a formato compatible con ReportLab"""
    try:
        ruta = buscar_imagen(ruta_img)
        if not ruta or not ruta.exists():
            return None
        
        img = Image.open(ruta)
        img = ImageOps.exif_transpose(img)
        img.thumbnail((450, 320))
        
        if img.mode == "RGBA":
            fondo = Image.new("RGB", img.size, "white")
            fondo.paste(img, mask=img.split()[3])
            img = fondo
        else:
            img = img.convert("RGB")
        
        nombre = f"pdf_{abs(hash(str(ruta)))}.png"
        salida = RUTA_TEMP / nombre
        img.save(salida, "PNG")
        return str(salida)
    except Exception:
        return None

# ============================================================
# CÁLCULO DE MÉTRICAS AVANZADAS
# ============================================================

def calcular_metricas_avanzadas(arq: Dict, capex: float, presupuesto_total: int, nodos: int = 1) -> Dict:
    """Calcular métricas completas incluyendo TCO, OPEX, PUE, ROI"""
    try:
        tdp = arq.get("tdp_total", 0)
        consumo = arq.get("consumo_real", 0)
        pue = arq.get("pue", 1.5)
        
        # Consumo energético
        consumo_anual_kwh = (consumo / 1000) * 24 * 365 * nodos * pue
        costo_kwh = 0.12
        opex_energia_anual = consumo_anual_kwh * costo_kwh
        
        # OPEX total (energía + mantenimiento + personal)
        mantenimiento_anual = capex * 0.15
        personal_anual = capex * 0.10
        opex_anual = opex_energia_anual + mantenimiento_anual + personal_anual
        
        # TCO a 5 años con inflación
        tco_5anos = capex + (opex_energia_anual * 5 * 1.08) + (mantenimiento_anual * 5) + (personal_anual * 5)
        
        # ROI
        roi_anios = capex / opex_anual if opex_anual > 0 else 0
        
        # Uso de presupuesto real
        costo_total = capex * nodos
        uso_presupuesto_pct = (costo_total / presupuesto_total * 100) if presupuesto_total > 0 else 0
        
        return {
            "tdp_por_nodo": tdp,
            "consumo_por_nodo_w": consumo,
            "consumo_total_w": consumo * nodos,
            "consumo_anual_kwh": consumo_anual_kwh,
            "costo_energia_anual_usd": opex_energia_anual,
            "opex_anual": opex_anual,
            "opex_mantenimiento": mantenimiento_anual,
            "opex_personal": personal_anual,
            "capex": capex,
            "tco_5anos": tco_5anos,
            "roi_anios": roi_anios,
            "uso_presupuesto_pct": uso_presupuesto_pct,
            "margen_disponible": presupuesto_total - costo_total,
            "semaforo_presupuesto": "✅ OK" if uso_presupuesto_pct <= 80 else "⚠️ AJUSTADO" if uso_presupuesto_pct <= 100 else "❌ EXCESO",
            "semaforo_energia": "✅ OK" if tdp <= 1000 else "⚠️ ALTO",
            "pue": pue,
        }
    except Exception as e:
        print(f"Error calculando métricas: {e}")
        return {}
    
def generar_diagnostico_experto(arq: Dict, metricas: Dict) -> str:
    """Generar diagnóstico técnico completo"""
    return f"""
╔═══════════════════════════════════════════════════════════════════╗
║               DIAGNÓSTICO TÉCNICO EXPERTO v2.1                    ║
╚═══════════════════════════════════════════════════════════════════╝

📊 PROCESAMIENTO:
  CPU: {arq.get('cpu_model', '?')} ({arq.get('cores', 0)} cores)
  Marca: {arq.get('cpu_brand', '?')}
  Arquitectura: Escalabilidad horizontal enterprise
  
💾 MEMORIA:
  RAM: {arq.get('ram_gb', 0)} GB
  Tipo: Configuración optimizada para throughput
  
🔄 ALMACENAMIENTO:
  Tipo: {arq.get('storage_type', '?')}
  Capacidad: {arq.get('storage_tb', 0)} TB
  Performance: {arq.get('iops', 0):,} IOPS
  Latencia estimada: < 100 µs (dependiendo de carga)

🖥️ ACELERACIÓN:
  GPU: {arq.get('gpu_model', 'Sin GPU')}
  VRAM: {arq.get('vram', 0)} GB
  TDP GPU: {arq.get('gpu_tdp', 0)} W
  
🌐 NETWORKING:
  Interfaz: {arq.get('nic_model', '?')}
  Velocity: {arq.get('nic_gbps', 0)} Gbps
  Latencia: {arq.get('latency_us', 0)} µs

❄️ ENFRIAMIENTO:
  Sistema: {arq.get('cool_type', '?')}
  PUE Estimado: {metricas.get('pue', 1.5):.2f}
  {'✅ Eficiente' if metricas.get('pue', 1.5) < 1.3 else '⚠️ Estándar'}

⚡ ENERGÍA:
  TDP Total: {metricas.get('tdp_por_nodo', 0)} W
  Consumo Real: {metricas.get('consumo_por_nodo_w', 0):.0f} W
  PSU: {arq.get('psu_watts', 0)} W (Eficiencia {arq.get('psu_eff', 0.9):.0%})
  
🔋 RESPALDO:
  UPS: {arq.get('ups_auto_min', 0)} minutos de autonomía
  {'✅ Suficiente' if arq.get('ups_auto_min', 0) >= 15 else '⚠️ Básico'}

💰 ECONOMÍA (Proyección 5 años):
  ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
  CAPEX:                    ${metricas.get('capex', 0):,.0f} USD
  ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
  OPEX Anual:
    • Energía:              ${metricas.get('costo_energia_anual_usd', 0):,.0f} USD/año
    • Mantenimiento (15%):  ${metricas.get('opex_mantenimiento', 0):,.0f} USD/año
    • Personal (10%):       ${metricas.get('opex_personal', 0):,.0f} USD/año
    ───────────────────────
    Total OPEX/año:         ${metricas.get('opex_anual', 0):,.0f} USD
  ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
  TCO (5 años):             ${metricas.get('tco_5anos', 0):,.0f} USD
  ROI:                      {metricas.get('roi_anios', 0):.1f} años
  Consumo Anual:            {metricas.get('consumo_anual_kwh', 0):,.0f} kWh
  
🎯 SEMÁFOROS DE VALIDACIÓN:
  Presupuesto: {metricas.get('semaforo_presupuesto', '?')}
  Energía: {metricas.get('semaforo_energia', '?')}
  Utilización: {metricas.get('uso_presupuesto_pct', 0):.1f}%

📋 RECOMENDACIONES TÉCNICAS:
  • Escalabilidad: {'Excelente para carga variable' if arq.get('cores', 0) >= 32 else 'Adecuada para cargas moderadas'}
  • Densidad: {arq.get('rack_u', 0)} unidades de rack
  • Eficiencia: PUE de {metricas.get('pue', 1.5):.2f} {'(por encima del promedio)' if metricas.get('pue', 1.5) < 1.4 else '(estándar industria)'}
  • TCO: {'Competitivo' if metricas.get('roi_anios', 0) < 3 else 'Estándar'} en relación CAPEX/OPEX
"""

# ============================================================
# NLP CON OLLAMA PHI-3
# ============================================================

def extraer_requisitos_nlp(texto: str) -> dict:
    """Extraer requisitos usando Phi-3 con detección avanzada"""
    prompt = """Eres un Arquitecto de Data Centers experto. Analiza el mensaje del usuario.

REGLAS DE INTENCIÓN:
- "saludo": Saluda (hola, buenos días, hey)
- "despedida": Se despide (adiós, gracias, bye)
- "ayuda": Pregunta qué haces o cómo funcionas
- "ajuste": Pide reducir, abaratar o buscar la opción más económica.
- "maximizar": Menciona un presupuesto alto, IA, HPC, o quiere el MEJOR equipo posible por su dinero.
- "tecnico": Da especificaciones técnicas normales.

REGLAS DE EXTRACCIÓN:
1. Convierte palabras a números ("ochenta mil" -> 80000)
2. cores: Número de núcleos de CPU
3. ram: GB de memoria RAM
4. iops: Operaciones I/O por segundo
5. gpu: 1 si menciona IA, inteligencia artificial, machine learning, gráficos
6. presupuesto: Cantidad en USD/dólares/dollars
7. Usa 0 si no se menciona

RESPONDE ÚNICAMENTE CON ESTE JSON:
{
  "intencion": "saludo|despedida|ayuda|ajuste|maximizar|tecnico",
  "mensaje_casual": "Tu respuesta empática y natural al usuario",
  "cores": 0, "ram": 0, "iops": 0, "gpu": 0, "presupuesto": 0
}

TEXTO DEL USUARIO: """ + texto
    
    try:
        resp = ollama.chat(
            model='phi3', 
            format='json', 
            messages=[{'role': 'user', 'content': prompt}],
            options={'num_ctx': 8192, 'temperature': 0.7}
        )
        
        contenido = resp['message']['content']
        print(f"\n--- RESPUESTA PHI-3 ---\n{contenido}\n-----------------------\n")
        
        # Extraer JSON del contenido
        match = re.search(r'\{.*?\}', contenido, re.DOTALL)
        if match:
            return json.loads(match.group(0))
        else:
            return json.loads(contenido.replace('```json', '').replace('```', '').strip())
    
    except Exception as e:
        print(f"❌ Error NLP: {e}")
        traceback.print_exc()
        return {
            "intencion": "error_ia",
            "mensaje_casual": f"Error conectando con Phi-3. Revisa que Ollama esté corriendo.",
            "cores": 0, "ram": 0, "iops": 0, "gpu": 0, "presupuesto": 0
        }

# ============================================================
# INTERFAZ GRÁFICA PRINCIPAL
# ============================================================

class SistemaExpertoDataCenter(ctk.CTk):
    """Aplicación principal con diseño visual premium"""
    
    def __init__(self):
        super().__init__()
        
        self.title("🏢 Sistema Experto de Data Centers v2.1")
        self.geometry("1650x950")
        self.minsize(1500, 850)
        
        # Estado
        self.motor = MotorProlog()
        self.contexto = ContextoUsuario()
        self.moneda = ctk.StringVar(value="USD")
        
        # Crear interfaz
        self._crear_interfaz()
        self._mensaje_bienvenida()
    
    def _crear_interfaz(self):
        """Construir toda la interfaz visual"""
        # Header
        header = ctk.CTkFrame(self, fg_color="#0a1628", height=90)
        header.pack(fill="x")
        header.pack_propagate(False)
        
        ctk.CTkLabel(
            header,
            text="🏢 Arquitecto Neuro-Simbólico IT - Sistema Experto v2.1",
            font=("Segoe UI", 28, "bold"),
            text_color=COLOR_INFO
        ).pack(pady=(18, 5))
        
        ctk.CTkLabel(
            header,
            text="Prolog v2.0 + NLP Phi-3 • Redundancia • Tier I-IV • TCO • Workloads • Ranking",
            font=("Segoe UI", 11),
            text_color="#888"
        ).pack(pady=(0, 12))
        
        # Contenedor principal
        main_frame = ctk.CTkFrame(self, fg_color="#111827")
        main_frame.pack(fill="both", expand=True, padx=12, pady=10)
        
        # Panel izquierdo: Chat
        self._crear_panel_chat(main_frame)
        
        # Panel derecho: Componentes
        self._crear_panel_componentes(main_frame)
        
        # Footer: Entrada y botones
        self._crear_footer()
    
    def _crear_panel_chat(self, parent):
        """Panel de chat con NLP"""
        left_frame = ctk.CTkFrame(parent, fg_color="#1a1a1a")
        left_frame.pack(side="left", fill="both", expand=True, padx=(0, 8))
        
        # Chat box
        self.chat_box = ctk.CTkTextbox(
            left_frame,
            font=("Consolas", 12),
            fg_color="#0d0d0d",
            text_color="#e0e0e0",
            wrap="word"
        )
        self.chat_box.pack(fill="both", expand=True, padx=10, pady=10)
        self.chat_box.configure(state="disabled")
        
        # Tags para colores
        self.chat_box.tag_config("usuario", foreground=COLOR_INFO)
        self.chat_box.tag_config("sistema", foreground=COLOR_DINERO)
        self.chat_box.tag_config("exito", foreground=COLOR_EXITO)
        self.chat_box.tag_config("error", foreground=COLOR_ERROR)
        self.chat_box.tag_config("advertencia", foreground=COLOR_ADVERTENCIA)
    
    def _crear_panel_componentes(self, parent):
        """Panel derecho con tarjetas de componentes"""
        right_frame = ctk.CTkFrame(parent, fg_color="#0d0d0d", width=750)
        right_frame.pack(side="right", fill="both", padx=(8, 0))
        right_frame.pack_propagate(False)
        
        ctk.CTkLabel(
            right_frame,
            text="🧩 Componentes Recomendados",
            font=("Segoe UI", 18, "bold"),
            text_color=COLOR_INFO
        ).pack(pady=12)
        
        # Frame scrollable para tarjetas
        self.componentes_frame = ctk.CTkScrollableFrame(right_frame, fg_color="#0d0d0d")
        self.componentes_frame.pack(fill="both", expand=True, padx=8, pady=(0, 8))
        
        # Mensaje inicial
        self.label_vacio = ctk.CTkLabel(
            self.componentes_frame,
            text="Sin arquitectura generada aún.\nUsa el chat para comenzar.",
            text_color="#888",
            font=("Segoe UI", 13)
        )
        self.label_vacio.pack(pady=50)
    
    def _crear_footer(self):
        """Footer con entrada, controles y botones"""
        footer = ctk.CTkFrame(self, fg_color="#0a1628")
        footer.pack(fill="x", padx=12, pady=10)
        
        # Entrada de texto
        entrada_frame = ctk.CTkFrame(footer, fg_color="#0a1628")
        entrada_frame.pack(fill="x", pady=(0, 8))
        
        self.entrada = ctk.CTkEntry(
            entrada_frame,
            placeholder_text="💬 Cuéntame qué necesitas (ej: 'servidor 64 cores, $50k, para IA')...",
            font=("Segoe UI", 13),
            height=42
        )
        self.entrada.pack(side="left", fill="both", expand=True, padx=(0, 8))
        self.entrada.bind("<Return>", lambda e: self._procesar_entrada())
        
        self.boton_enviar = ctk.CTkButton(
            entrada_frame,
            text="📤 Enviar",
            command=self._procesar_entrada,
            width=100,
            height=42,
            font=("Segoe UI", 12, "bold"),
            fg_color=COLOR_INFO,
            hover_color="#0088ff"
        )
        self.boton_enviar.pack(side="right")
        
        # Controles avanzados
        controles_frame = ctk.CTkFrame(footer, fg_color="#0a1628")
        controles_frame.pack(fill="x", pady=(0, 8))
        
        # Selector de moneda
        ctk.CTkLabel(controles_frame, text="💱", font=("Segoe UI", 12)).pack(side="left", padx=(4, 2))
        selector_moneda = ctk.CTkOptionMenu(
            controles_frame,
            values=["USD", "MXN"],
            variable=self.moneda,
            command=self._refrescar_moneda,
            width=80,
            font=("Segoe UI", 11)
        )
        selector_moneda.pack(side="left", padx=(0, 12))
        
        # Selector de redundancia
        ctk.CTkLabel(controles_frame, text="🔄 Red:", font=("Segoe UI", 11)).pack(side="left", padx=(0, 4))
        self.red_var = ctk.StringVar(value="ninguna")
        red_menu = ctk.CTkOptionMenu(
            controles_frame,
            values=["ninguna", "n_mas_1", "n_mas_2", "dos_n"],
            variable=self.red_var,
            width=100,
            font=("Segoe UI", 11)
        )
        red_menu.pack(side="left", padx=(0, 12))
        
        # Selector de Tier
        ctk.CTkLabel(controles_frame, text="📊 Tier:", font=("Segoe UI", 11)).pack(side="left", padx=(0, 4))
        self.tier_var = ctk.StringVar(value="tier3")
        tier_menu = ctk.CTkOptionMenu(
            controles_frame,
            values=["tier1", "tier2", "tier3", "tier4"],
            variable=self.tier_var,
            width=90,
            font=("Segoe UI", 11)
        )
        tier_menu.pack(side="left", padx=(0, 12))
        
        # Años para TCO
        ctk.CTkLabel(controles_frame, text="📅 Años TCO:", font=("Segoe UI", 11)).pack(side="left", padx=(0, 4))
        self.anios_var = ctk.StringVar(value="5")
        anios_menu = ctk.CTkOptionMenu(
            controles_frame,
            values=["3", "5", "7", "10"],
            variable=self.anios_var,
            width=70,
            font=("Segoe UI", 11)
        )
        anios_menu.pack(side="left")
        
        # Botones de acción
        botones_frame = ctk.CTkFrame(footer, fg_color="#0a1628")
        botones_frame.pack(fill="x", pady=(0, 8))
        
        botones = [
            ("🏆 Top-5", self._mostrar_ranking),
            ("🧮 Comparar", self._comparar_arquitecturas),
            ("🧠 Diagnóstico", self._mostrar_diagnostico),
            ("📄 PDF", self._generar_pdf),
            ("📈 Gráfica", self._generar_grafica),
            ("📊 Excel", self._exportar_excel),
            ("💾 JSON", self._guardar_json),
            ("🆕 Nuevo", self._nuevo_diseno),
        ]
        
        for texto, comando in botones:
            ctk.CTkButton(
                botones_frame,
                text=texto,
                command=comando,
                width=100,
                height=32,
                font=("Segoe UI", 11),
                fg_color="#1f2937",
                hover_color="#374151"
            ).pack(side="left", padx=4)
        
        # Barra de progreso
        self.barra_progreso = ctk.CTkProgressBar(footer, mode="indeterminate")
        self.barra_progreso.pack(fill="x", pady=(8, 0))
        self.barra_progreso.set(0)
    
    # ========== MÉTODOS DE CHAT ==========
    
    def _agregar_mensaje(self, rol: str, texto: str, etiqueta: str = ""):
        """Agregar mensaje al chat"""
        self.chat_box.configure(state="normal")
        
        if rol == "usuario":
            self.chat_box.insert("end", f"\n👤 TÚ:\n", "usuario")
            self.chat_box.insert("end", f"{texto}\n")
        elif rol == "sistema":
            self.chat_box.insert("end", f"\n🤖 SISTEMA:\n", "sistema")
            self.chat_box.insert("end", f"{texto}\n", etiqueta if etiqueta else "sistema")
        
        self.chat_box.configure(state="disabled")
        self.chat_box.see("end")
    
    def _mensaje_bienvenida(self):
        """Mostrar mensaje de bienvenida"""
        bienvenida = """🚀 SISTEMA EXPERTO NEURO-SIMBÓLICO INICIADO

Soy tu Arquitecto Virtual especializado en infraestructura de Data Centers.

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
CÓMO USAR:
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

1️⃣ HABLA NATURAL: Dime qué necesitas en lenguaje coloquial
   • "Necesito un servidor web con 32 cores, tengo $20,000"
   • "Diseña un cluster para IA, presupuesto $80k"
   • "Está muy caro, ¿puedes ajustarlo a mi presupuesto?"

2️⃣ AJUSTA CONFIGURACIÓN: Usa los selectores de abajo
   • Moneda: USD / MXN
   • Redundancia: ninguna / N+1 / N+2 / 2N
   • Tier: I, II, III, IV (certificación Uptime Institute)
   • Años TCO: 3, 5, 7, 10 (proyección financiera)

3️⃣ EXPLORA OPCIONES:
   • Top-5: Ranking de mejores arquitecturas
   • Comparar: 3 opciones (Económica/Balanceada/Premium)
   • Diagnóstico: Análisis técnico completo con TCO
   • Exporta: PDF profesional, Excel, Gráficas, JSON

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
NUEVAS CARACTERÍSTICAS v2.1:
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

✅ 10 componentes visuales (CPU, MB, RAM, Storage, GPU, NIC, Cooling, PSU, UPS, Chasis)
✅ TCO real con CAPEX + OPEX proyectado (energía + mantenimiento + personal)
✅ Redundancia N+1, N+2, 2N
✅ Certificación Tier I-IV (Uptime Institute)
✅ PUE y eficiencia energética
✅ Workload profiles (AI, HPC, Database, Web, etc.)
✅ Scoring multicriterio
✅ Compatibilidad PCIe CPU-GPU
✅ Diagnóstico avanzado con ROI

¡Adelante, cuéntame qué necesitas!
"""
        self._agregar_mensaje("sistema", bienvenida)
    
    def _procesar_entrada(self):
        """Procesar entrada del usuario"""
        texto = self.entrada.get().strip()
        if not texto:
            return
        
        self._agregar_mensaje("usuario", texto)
        self.entrada.delete(0, "end")
        self.boton_enviar.configure(state="disabled")
        self.barra_progreso.start()
        
        # Ejecutar en hilo separado
        threading.Thread(target=self._procesar_en_background, args=(texto,), daemon=True).start()
    
    def _procesar_en_background(self, texto: str):
        """Procesar consulta en segundo plano"""
        try:
            # Extraer requisitos con NLP
            req = extraer_requisitos_nlp(texto)
            intencion = req.get("intencion", "tecnico")
            msg_casual = req.get("mensaje_casual", "")
            
            # Manejo de intenciones
            if intencion == "error_ia":
                self._agregar_mensaje("sistema", msg_casual, "error")
                return
            
            if intencion in ["saludo", "ayuda", "despedida"]:
                self._agregar_mensaje("sistema", msg_casual, "info")
                if intencion == "despedida":
                    self.contexto.limpiar()
                return
            
            # Extraer valores numéricos
            try:
                cores = int(req.get('cores', 0))
                ram = int(req.get('ram', 0))
                iops = int(req.get('iops', 0))
                gpu = int(req.get('gpu', 0))
                presupuesto = int(req.get('presupuesto', 0))
            except:
                cores = ram = iops = gpu = presupuesto = 0
            
            # --- 1. DETECCIÓN DE WORKLOAD Y AJUSTE DE TIER ---
            # Mapeo de palabras clave en el texto hacia los workloads de tu Prolog
            workload_map = {
                "ia": "ai_training", "inteligencia artificial": "ai_training",
                "pago": "database_oltp", "transaccional": "database_oltp",
                "hpc": "hpc_simulation", "simulacion": "hpc_simulation",
                "web": "web_hosting", "virtual": "virtualizacion"
            }
            
            workload_atom = None
            for key, atom in workload_map.items():
                if key in texto.lower():
                    workload_atom = atom
                    break
            
            if workload_atom:
                # Consultamos al motor Prolog qué nivel de Tier es el mínimo obligatorio
                res_tier = list(self.motor.prolog.query(f"min_tier_workload({workload_atom}, T)"))
                if res_tier:
                    tier_min = res_tier[0]['T']
                    tier_ui = self.tier_var.get().lower()
                    valores = {'tier1': 1, 'tier2': 2, 'tier3': 3, 'tier4': 4}
                    
                    if valores.get(tier_ui, 1) < valores.get(tier_min, 1):
                        self._agregar_mensaje("sistema", 
                            f"⚠️ Sugerencia: Para '{workload_atom.replace('_', ' ')}', el sistema requiere al menos {tier_min.upper()}. "
                            f"He ajustado el selector de Tier automáticamente.", "advertencia")
                        self.tier_var.set(tier_min)

            # --- 2. DETECCIÓN AUTOMÁTICA DE ALTO RENDIMIENTO ---
            if intencion == "tecnico" and (presupuesto >= 30000 or gpu == 1):
                intencion = "maximizar"
                msg_casual = "Detecté un requerimiento de alto rendimiento (HPC/IA). Voy a exprimir ese presupuesto al máximo."
            
            # Actualizar contexto
            if intencion == "ajuste":
                self.contexto.cores_min = 0
                self.contexto.ram_min = 0
                self.contexto.iops_min = 0
                self._agregar_mensaje("sistema", f"🔄 {msg_casual}\nBuscando la opción más económica...", "info")
            elif intencion == "maximizar":
                self._agregar_mensaje("sistema", f"🚀 {msg_casual}\nBuscando la arquitectura más potente posible...", "info")
            
            if cores > 0: self.contexto.cores_min = cores
            if ram > 0: self.contexto.ram_min = ram
            if iops > 0: self.contexto.iops_min = iops
            if gpu > 0: self.contexto.gpu_vram_min = 1
            if presupuesto > 0: self.contexto.presupuesto = presupuesto
            
            # Validar que tengamos suficiente información
            if self.contexto.presupuesto > 0 and (
                self.contexto.cores_min > 0 or 
                self.contexto.ram_min > 0 or 
                intencion in ["ajuste", "maximizar"]
            ):
                if msg_casual and intencion not in ["ajuste", "maximizar"]:
                    self._agregar_mensaje("sistema", f"💬 {msg_casual}\n\n⚙️ Consultando motor Prolog v2.0...", "info")
                
                self._ejecutar_consulta_prolog(intencion)
            else:
                aviso = "❓ Necesito más información: Presupuesto y Cores/RAM mínimos."
                self._agregar_mensaje("sistema", aviso, "advertencia")
        
        except Exception as e:
            self._agregar_mensaje("sistema", f"❌ Error interno: {str(e)}", "error")
            traceback.print_exc()
        
        finally:
            self.barra_progreso.stop()
            self.boton_enviar.configure(state="normal")
    
    # SE RECIBE LA INTENCIÓN
    def _ejecutar_consulta_prolog(self, intencion="tecnico"):
        """Ejecutar consulta a Prolog con parámetros avanzados"""
        try:
            if not self.motor.cargado:
                self._agregar_mensaje("sistema", "❌ Motor Prolog no cargado. Verifica el archivo .pl", "error")
                return
            
            # Obtener parámetros de los controles
            redundancia = self.red_var.get()
            anios = int(self.anios_var.get())
            tier_seleccionado = self.tier_var.get()  # <-- CAPTURAMOS EL TIER DE LA INTERFAZ
            
            # Consulta arquitectura completa
            resultados_brutos = self.motor.arquitectura_completa(
                self.contexto.cores_min,
                self.contexto.ram_min,
                self.contexto.iops_min,
                self.contexto.gpu_vram_min,
                self.contexto.presupuesto,
                redundancia,
                anios
            )
            
            # <-- APLICAMOS EL FILTRO DEL TIER EXACTO QUE PEDISTE -->
            resultados = [r for r in resultados_brutos if str(r.get("Tier", "")).lower() == tier_seleccionado.lower()]
            
            if resultados:
                # SE MODIFICA EL ORDENAMIENTO EN BASE A LA INTENCIÓN
                if intencion == "maximizar":
                    resultados.sort(key=lambda x: x.get("CAPEX", 0), reverse=True)
                else:
                    resultados.sort(key=lambda x: x.get("CAPEX", float('inf')))
                    
                self._mostrar_arquitectura(resultados[0])
            else:
                # No hay resultados, mostrar diagnóstico
                self._agregar_mensaje("sistema",
                    f"⚠️ No se encontró arquitectura válida para {tier_seleccionado.upper()}.\n\n"
                    "💡 OPCIONES:\n"
                    "• Aumentar presupuesto\n"
                    "• Bajar el nivel de Tier\n"
                    "• Cambiar redundancia a 'ninguna'\n\n"
                    "🔍 Ejecutando diagnóstico automático...", "advertencia")
                
                # Ejecutar diagnóstico
                self.motor.diagnostico(
                    self.contexto.cores_min,
                    self.contexto.ram_min,
                    self.contexto.iops_min,
                    self.contexto.gpu_vram_min,
                    self.contexto.presupuesto
                )
                self._agregar_mensaje("sistema", "Revisa la consola para ver el diagnóstico detallado.")
        
        except Exception as e:
            self._agregar_mensaje("sistema", f"❌ Error en consulta Prolog: {str(e)}", "error")
            traceback.print_exc()
    
    def _mostrar_arquitectura(self, resultado: Dict):
        """Mostrar arquitectura en panel de componentes y chat"""
        try:
            arq_str = str(resultado.get("Arq", ""))
            arq = parsear_arquitectura_41_campos(arq_str)
            
            if not arq:
                self._agregar_mensaje("sistema", "❌ Error parseando arquitectura", "error")
                return
            
            # Calcular métricas
            capex = resultado.get("CAPEX", self.contexto.presupuesto)
            metricas = calcular_metricas_avanzadas(arq, capex, self.contexto.presupuesto)
            metricas["capex"] = capex
            
            # TCO desde Prolog
            tco_prolog = resultado.get("TCO", None)
            if tco_prolog and isinstance(tco_prolog, (int, float)):
                metricas["tco_5anos"] = tco_prolog
            
            # Tier alcanzado
            tier = resultado.get("Tier", "tier2")
            
            # Guardar en contexto
            self.contexto.ultima_arquitectura = {
                "arquitectura": arq,
                "metricas": metricas,
                "tier": tier,
                "resultado": resultado,
                "timestamp": datetime.now().isoformat()
            }
            
            # Mostrar en panel de tarjetas
            self._renderizar_tarjetas(arq, metricas)
            
            # Mensaje en chat
            moneda = self.moneda.get()
            msg = f"""✅ ARQUITECTURA DISEÑADA EXITOSAMENTE

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
📊 RESUMEN TÉCNICO:
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
  CPU:      {arq.get('cpu_model', '?')} ({arq.get('cores', 0)} cores)
  RAM:      {arq.get('ram_gb', 0)} GB
  Storage:  {arq.get('iops', 0):,} IOPS
  GPU:      {arq.get('gpu_model', 'Sin GPU')}
  Red:      {arq.get('nic_gbps', 0)} Gbps

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
💰 ECONOMÍA:
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
  CAPEX:           {formatear_precio(metricas.get('capex', 0), moneda)}
  OPEX/año:        {formatear_precio(metricas.get('opex_anual', 0), moneda)}
  TCO 5 años:      {formatear_precio(metricas.get('tco_5anos', 0), moneda)}
  ROI:             {metricas.get('roi_anios', 0):.1f} años

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
⚡ EFICIENCIA:
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
  Consumo:         {metricas.get('consumo_por_nodo_w', 0):.0f} W/nodo
  PUE:             {metricas.get('pue', 1.5):.2f}
  Tier:            {tier.upper()}
  Redundancia:     {self.red_var.get().upper().replace('_MAS_', '+').replace('DOS_N', '2N')}
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
🎯 ESTADO:
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
  Presupuesto:     {metricas.get('semaforo_presupuesto', '?')}
  Uso:             {metricas.get('uso_presupuesto_pct', 0):.1f}%
  Margen:          {formatear_precio(metricas.get('margen_disponible', 0), moneda)}

📋 Ahora puedes usar:
   • Top-5: Ver ranking completo
   • Comparar: 3 opciones (Económica/Balanceada/Premium)
   • Diagnóstico: Análisis técnico detallado
   • Exportar: PDF, Excel, Gráficas, JSON
"""
            self._agregar_mensaje("sistema", msg, "exito")
        
        except Exception as e:
            self._agregar_mensaje("sistema", f"❌ Error mostrando arquitectura: {str(e)}", "error")
            traceback.print_exc()
    
    def _renderizar_tarjetas(self, arq: Dict, metricas: Dict):
        """Renderizar tarjetas visuales de los 10 componentes"""
        # Limpiar panel
        for widget in self.componentes_frame.winfo_children():
            widget.destroy()
        
        moneda = self.moneda.get()
        
        # Definir los 10 componentes con sus datos
        componentes = [
            ("💻 CPU", f"{arq.get('cpu_model', '?')}\n{arq.get('cores', 0)} cores", 
             arq.get('img_cpu'), f"${metricas.get('capex', 0) * 0.30:,.0f}"),
            
            ("🔌 Placa Base", arq.get('mb_id', '?'), 
             arq.get('img_mb'), f"${metricas.get('capex', 0) * 0.10:,.0f}"),
            
            ("🧠 Memoria RAM", f"{arq.get('ram_gb', 0)} GB", 
             arq.get('img_ram'), f"${metricas.get('capex', 0) * 0.15:,.0f}"),
            
            ("💾 Almacenamiento", f"{arq.get('storage_type', '?')}\n{arq.get('iops', 0):,} IOPS", 
             arq.get('img_storage'), f"${metricas.get('capex', 0) * 0.08:,.0f}"),
            
            ("🖥️ GPU", f"{arq.get('gpu_model', 'Sin GPU')}\n{arq.get('vram', 0)} GB VRAM", 
             arq.get('img_gpu'), f"${metricas.get('capex', 0) * 0.20:,.0f}"),
            
            ("🌐 Red (NIC)", f"{arq.get('nic_model', '?')}\n{arq.get('nic_gbps', 0)} Gbps", 
             arq.get('img_nic'), f"${metricas.get('capex', 0) * 0.05:,.0f}"),
            
            ("❄️ Enfriamiento", f"{arq.get('cool_type', '?')}\nPUE {arq.get('pue', 1.5):.2f}", 
             arq.get('img_cool'), f"${metricas.get('capex', 0) * 0.04:,.0f}"),
            
            ("⚡ PSU", f"{arq.get('psu_watts', 0)} W\n{arq.get('psu_eff', 0.9):.0%} Eficiencia", 
             arq.get('img_psu'), f"${metricas.get('capex', 0) * 0.04:,.0f}"),
            
            ("🔋 UPS", f"{arq.get('ups_auto_min', 0)} min autonomía", 
             arq.get('img_ups'), f"${metricas.get('capex', 0) * 0.03:,.0f}"),
            
            ("🏗️ Chasis", f"{arq.get('rack_u', 0)} Unidades Rack", 
             arq.get('img_chasis'), f"${metricas.get('capex', 0) * 0.01:,.0f}"),
        ]
        
        # Crear tarjetas en filas de 2
        for i in range(0, len(componentes), 2):
            fila = ctk.CTkFrame(self.componentes_frame, fg_color="transparent")
            fila.pack(fill="x", padx=8, pady=6)
            
            # Primera tarjeta
            self._crear_tarjeta(fila, *componentes[i])
            
            # Segunda tarjeta si existe
            if i + 1 < len(componentes):
                self._crear_tarjeta(fila, *componentes[i + 1])
    
    def _crear_tarjeta(self, parent, titulo: str, nombre: str, img_path: str, precio: str):
        """Crear tarjeta visual individual"""
        card = ctk.CTkFrame(parent, fg_color="#111827", corner_radius=18, width=335, height=240)
        card.pack(side="left", padx=8, pady=4)
        card.pack_propagate(False)
        
        # Header
        header = ctk.CTkFrame(card, fg_color="#0a1628", height=40, corner_radius=12)
        header.pack(fill="x", padx=0, pady=(0, 6))
        header.pack_propagate(False)
        
        ctk.CTkLabel(
            header,
            text=titulo,
            font=("Segoe UI", 14, "bold"),
            text_color=COLOR_INFO
        ).pack(expand=True)
        
        # Imagen
        img_box = ctk.CTkFrame(card, fg_color="#ffffff", corner_radius=8, width=300, height=110)
        img_box.pack(pady=(4, 8))
        img_box.pack_propagate(False)
        
        ruta_img = buscar_imagen(img_path) if img_path else None
        img_ctk = crear_imagen_ctk(ruta_img, tamaño=(280, 100))
        
        if img_ctk:
            img_label = ctk.CTkLabel(img_box, text="", image=img_ctk, width=280)
            img_label.pack(expand=True)
        else:
            ctk.CTkLabel(
                img_box,
                text="📦 Sin imagen",
                text_color="#999",
                font=("Segoe UI", 12, "bold")
            ).pack(expand=True)
        
        # Nombre
        ctk.CTkLabel(
            card,
            text=nombre,
            wraplength=310,
            font=("Segoe UI", 12, "bold"),
            text_color="#ffffff",
            justify="center"
        ).pack(pady=(2, 4))
        
        # Precio
        ctk.CTkLabel(
            card,
            text=precio,
            text_color=COLOR_DINERO,
            font=("Segoe UI", 13, "bold")
        ).pack(pady=(0, 8))
    
    # ========== FUNCIONES DE BOTONES ==========
    
    def _mostrar_ranking(self):
        """Mostrar Top-5 arquitecturas"""
        if not self.motor.cargado:
            self._agregar_mensaje("sistema", "❌ Motor Prolog no disponible", "error")
            return
        
        try:
            self._agregar_mensaje("sistema", "🔍 Buscando Top-5 arquitecturas...", "info")
            
            resultados = self.motor.consulta_ranking(
                5,
                self.contexto.cores_min,
                self.contexto.ram_min,
                self.contexto.iops_min,
                self.contexto.gpu_vram_min,
                self.contexto.presupuesto
            )
            
            if resultados and len(resultados) > 0:
                msg = "🏆 TOP-5 ARQUITECTURAS (por score):\n\n"
                
                # Extraer TopN del resultado
                top_lista = resultados[0].get('TopN', [])
                
                for i, item in enumerate(top_lista, 1):
                    # Parsear el item
                    score_parte = str(item).split('-')[0]
                    arq_str = str(item).split('arq_rankeada(')[1].rstrip(')')
                    
                    try:
                        score = float(score_parte)
                    except:
                        score = 0.0
                    
                    partes_arq = arq_str.split(',', 1)
                    costo = int(partes_arq[0]) if len(partes_arq) > 0 else 0
                    
                    # Parsear arquitectura completa
                    arq_completa_str = partes_arq[1] if len(partes_arq) > 1 else ""
                    arq_data = parsear_arquitectura_41_campos(f"arq({arq_completa_str}")
                    
                    msg += f"{i}. {arq_data.get('cpu_model', '?')} - "
                    msg += f"{formatear_precio(costo, self.moneda.get())} "
                    msg += f"(Score: {score:.1f})\n"
                
                self._agregar_mensaje("sistema", msg, "exito")
            else:
                self._agregar_mensaje("sistema", "⚠️ No hay suficientes opciones para ranking", "advertencia")
        
        except Exception as e:
            self._agregar_mensaje("sistema", f"❌ Error en ranking: {str(e)}", "error")
            traceback.print_exc()
    
    def _comparar_arquitecturas(self):
        """Comparador visual de 3 opciones (Económica/Balanceada/Premium)"""
        if self.contexto.presupuesto <= 0:
            self._agregar_mensaje("sistema", "⚠️ Primero genera una arquitectura", "advertencia")
            return
        
        try:
            # Obtener Top-3
            resultados = self.motor.consulta_ranking(
                3,
                self.contexto.cores_min if self.contexto.cores_min > 0 else 8,
                self.contexto.ram_min if self.contexto.ram_min > 0 else 16,
                self.contexto.iops_min if self.contexto.iops_min > 0 else 1000,
                self.contexto.gpu_vram_min,
                self.contexto.presupuesto
            )
            
            if not resultados or len(resultados) == 0:
                self._agregar_mensaje("sistema", "⚠️ No hay opciones para comparar", "advertencia")
                return
            
            top_lista = resultados[0].get('TopN', [])
            if len(top_lista) < 2:
                self._agregar_mensaje("sistema", "⚠️ Se necesitan al menos 2 opciones para comparar", "advertencia")
                return
            
            # Parsear las 3 opciones
            opciones = []
            for item in top_lista[:3]:
                try:
                    arq_str = str(item).split('arq_rankeada(')[1].rstrip(')')
                    partes = arq_str.split(',', 1)
                    costo = int(partes[0]) if len(partes) > 0 else 0
                    arq_completa = parsear_arquitectura_41_campos(f"arq({partes[1]}")
                    
                    opciones.append({
                        'costo': costo,
                        'arq': arq_completa
                    })
                except:
                    continue
            
            if len(opciones) < 2:
                self._agregar_mensaje("sistema", "⚠️ Error parseando opciones", "error")
                return
            
            # Crear ventana comparativa
            ventana = ctk.CTkToplevel(self)
            ventana.title("🧮 Comparador Visual de Arquitecturas")
            ventana.geometry("1400x700")
            ventana.minsize(1200, 600)
            ventana.configure(fg_color="#0b1220")
            
            # Header
            ctk.CTkLabel(
                ventana,
                text="📊 Comparador de Arquitecturas",
                font=("Segoe UI", 28, "bold"),
                text_color=COLOR_INFO
            ).pack(pady=(20, 8))
            
            ctk.CTkLabel(
                ventana,
                text=f"Moneda: {self.moneda.get()}",
                font=("Segoe UI", 13, "bold"),
                text_color=COLOR_DINERO
            ).pack(pady=(0, 12))
            
            # Contenedor de opciones
            contenedor = ctk.CTkFrame(ventana, fg_color="#111827")
            contenedor.pack(fill="both", expand=True, padx=20, pady=15)
            
            nombres = ["💸 ECONÓMICA", "⚖️ BALANCEADA", "💎 PREMIUM"]
            colores = [COLOR_EXITO, "#38bdf8", COLOR_ADVERTENCIA]
            
            for i, opcion in enumerate(opciones):
                card = ctk.CTkFrame(contenedor, fg_color="#0f172a", corner_radius=22)
                card.pack(side="left", fill="both", expand=True, padx=12, pady=12)
                
                # Título
                ctk.CTkLabel(
                    card,
                    text=nombres[i] if i < len(nombres) else f"OPCIÓN {i+1}",
                    font=("Segoe UI", 22, "bold"),
                    text_color=colores[i] if i < len(colores) else COLOR_INFO
                ).pack(pady=(18, 8))
                
                # Precio
                ctk.CTkLabel(
                    card,
                    text=formatear_precio(opcion['costo'], self.moneda.get()),
                    font=("Segoe UI", 24, "bold"),
                    text_color=COLOR_DINERO
                ).pack(pady=(0, 10))
                
                # Tabla de specs
                tabla = ctk.CTkFrame(card, fg_color="#111827", corner_radius=14)
                tabla.pack(fill="both", expand=True, padx=14, pady=12)
                
                arq = opcion['arq']
                specs = [
                    ("CPU", f"{arq.get('cpu_model', '?')} ({arq.get('cores', 0)} cores)"),
                    ("RAM", f"{arq.get('ram_gb', 0)} GB"),
                    ("Storage", f"{arq.get('iops', 0):,} IOPS"),
                    ("GPU", arq.get('gpu_model', 'Sin GPU')),
                    ("Red", f"{arq.get('nic_gbps', 0)} Gbps"),
                    ("PUE", f"{arq.get('pue', 1.5):.2f}"),
                ]
                
                for label, valor in specs:
                    fila = ctk.CTkFrame(tabla, fg_color="transparent")
                    fila.pack(fill="x", padx=12, pady=6)
                    
                    ctk.CTkLabel(
                        fila,
                        text=f"{label}:",
                        width=80,
                        anchor="w",
                        font=("Segoe UI", 12, "bold"),
                        text_color=COLOR_INFO
                    ).pack(side="left")
                    
                    ctk.CTkLabel(
                        fila,
                        text=str(valor),
                        anchor="w",
                        justify="left",
                        wraplength=220,
                        font=("Segoe UI", 11),
                        text_color="#ffffff"
                    ).pack(side="left", fill="x", expand=True)
            
            self._agregar_mensaje("sistema", "🧮 Comparador visual abierto", "exito")
        
        except Exception as e:
            self._agregar_mensaje("sistema", f"❌ Error en comparador: {str(e)}", "error")
            traceback.print_exc()
    
    def _mostrar_diagnostico(self):
        """Mostrar diagnóstico técnico completo"""
        if not self.contexto.ultima_arquitectura:
            self._agregar_mensaje("sistema", "❌ Primero genera una arquitectura", "error")
            return
        
        arq = self.contexto.ultima_arquitectura.get("arquitectura", {})
        metricas = self.contexto.ultima_arquitectura.get("metricas", {})
        
        diagnostico = generar_diagnostico_experto(arq, metricas)
        self._agregar_mensaje("sistema", diagnostico)
    
    # SE INTEGRA LA CORRECCIÓN DEL PDF CON LA TABLA COMPLETA E IMÁGENES A DOBLE COLUMNA
    def _generar_pdf(self):
        """Generar reporte PDF profesional"""
        if not REPORTLAB_AVAILABLE:
            self._agregar_mensaje("sistema", "⚠️ Instala reportlab: pip install reportlab", "advertencia")
            return
        
        if not self.contexto.ultima_arquitectura:
            self._agregar_mensaje("sistema", "❌ Primero genera una arquitectura", "error")
            return
        
        try:
            arq = self.contexto.ultima_arquitectura.get("arquitectura", {})
            metricas = self.contexto.ultima_arquitectura.get("metricas", {})
            
            fecha = datetime.now().strftime("%Y%m%d_%H%M%S")
            ruta_pdf = RUTA_REPORTES / f"datacenter_v2_{fecha}.pdf"
            
            doc = SimpleDocTemplate(str(ruta_pdf), pagesize=letter, rightMargin=35, leftMargin=35, topMargin=35, bottomMargin=35)
            story = []
            styles = getSampleStyleSheet()
            
            # Título
            titulo_style = ParagraphStyle(
                "CustomTitle",
                parent=styles["Title"],
                fontSize=20,
                textColor=colors.HexColor("#00AEEF"),
                spaceAfter=14
            )
            story.append(Paragraph("Reporte Profesional de Data Center v2.1", titulo_style))
            story.append(Paragraph("Sistema Experto Neuro-Simbólico", styles["Normal"]))
            story.append(Spacer(1, 14))
            
            # Tabla resumen con todos los componentes
            datos_resumen = [
                ["Campo", "Valor"],
                ["Fecha", datetime.now().strftime("%d/%m/%Y %H:%M")],
                ["CPU", f"{arq.get('cpu_model', '?')} ({arq.get('cores', 0)} cores)"],
                ["Placa Base", arq.get('mb_id', '?')],
                ["RAM", f"{arq.get('ram_gb', 0)} GB"],
                ["Almacenamiento", f"{arq.get('storage_type', '?')} - {arq.get('iops', 0):,} IOPS"],
                ["GPU", arq.get('gpu_model', 'Sin GPU')],
                ["Red (NIC)", f"{arq.get('nic_model', '?')} ({arq.get('nic_gbps', 0)} Gbps)"],
                ["Enfriamiento", arq.get('cool_type', '?')],
                ["PSU", f"{arq.get('psu_watts', 0)} W"],
                ["UPS", f"{arq.get('ups_auto_min', 0)} min"],
                ["Chasis", f"{arq.get('rack_u', 0)} U"],
                ["CAPEX", formatear_precio(metricas.get('capex', 0))],
                ["OPEX/año", formatear_precio(metricas.get('opex_anual', 0))],
                ["TCO 5 años", formatear_precio(metricas.get('tco_5anos', 0))],
                ["PUE", f"{metricas.get('pue', 1.5):.2f}"],
                ["ROI", f"{metricas.get('roi_anios', 0):.1f} años"],
            ]
            
            tabla = Table(datos_resumen, colWidths=[2.5*inch, 4.35*inch])
            tabla.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (1, 0), colors.HexColor("#0A1628")),
                ("TEXTCOLOR", (0, 0), (1, 0), colors.whitesmoke),
                ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 10),
                ("GRID", (0, 0), (-1, -1), 1, colors.HexColor("#CBD5E1")),
                ("PADDING", (0, 0), (-1, -1), 8),
            ]))
            
            story.append(tabla)
            story.append(Spacer(1, 16))
            
            # Agregar imágenes de LOS 10 componentes
            subtitulo_style = ParagraphStyle("Subtitulo", parent=styles["Heading2"], fontSize=14, textColor=colors.HexColor("#0A1628"), spaceBefore=12, spaceAfter=10)
            story.append(Paragraph("Componentes Principales", subtitulo_style))
            
            componentes_img = [
                ("CPU", arq.get('img_cpu')),
                ("Placa Base", arq.get('img_mb')),
                ("Memoria RAM", arq.get('img_ram')),
                ("Almacenamiento", arq.get('img_storage')),
                ("GPU", arq.get('img_gpu')),
                ("Red (NIC)", arq.get('img_nic')),
                ("Enfriamiento", arq.get('img_cool')),
                ("PSU", arq.get('img_psu')),
                ("UPS", arq.get('img_ups')),
                ("Chasis", arq.get('img_chasis')),
            ]
            
            # Formato de cuadrícula en 2 columnas para optimizar espacio en el PDF
            img_table_data = []
            row = []
            for nombre, img_path in componentes_img:
                img_preparada = preparar_imagen_pdf(img_path) if img_path else None
                if img_preparada:
                    try:
                        celda = [
                            Paragraph(f"<b>{nombre}</b>", styles["Normal"]),
                            Spacer(1, 4),
                            RLImage(img_preparada, width=2.5*inch, height=1.6*inch)
                        ]
                        row.append(celda)
                        if len(row) == 2:
                            img_table_data.append(row)
                            row = []
                    except:
                        pass
            
            if row: # Rellenar espacio si hay número impar de imágenes
                row.append("")
                img_table_data.append(row)

            if img_table_data:
                img_table = Table(img_table_data, colWidths=[3.5*inch, 3.5*inch])
                img_table.setStyle(TableStyle([
                    ('VALIGN', (0,0), (-1,-1), 'TOP'),
                    ('ALIGN', (0,0), (-1,-1), 'LEFT'),
                    ('BOTTOMPADDING', (0,0), (-1,-1), 15),
                ]))
                story.append(img_table)
            
            doc.build(story)
            
            # Abrir PDF
            if os.name == "nt":
                os.startfile(str(ruta_pdf))
            else:
                os.system(f"open {ruta_pdf}")
            
            self._agregar_mensaje("sistema", f"✅ PDF generado: {ruta_pdf.name}", "exito")
        
        except Exception as e:
            self._agregar_mensaje("sistema", f"❌ Error generando PDF: {e}", "error")
            traceback.print_exc()
    
    def _generar_grafica(self):
        """Generar gráfica de presupuesto y consumo"""
        if not self.contexto.ultima_arquitectura:
            self._agregar_mensaje("sistema", "❌ Primero genera una arquitectura", "error")
            return
        
        try:
            metricas = self.contexto.ultima_arquitectura.get("metricas", {})
            capex = metricas.get("capex", 0)
            margen = metricas.get("margen_disponible", 0)
            opex_anual = metricas.get("opex_anual", 0)
            
            fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(14, 6))
            
            # Gráfica 1: Presupuesto
            ax1.bar(["Usado", "Disponible"], [capex, max(margen, 0)], color=["#00AEEF", COLOR_EXITO])
            ax1.set_title("Análisis de Presupuesto", fontsize=14, fontweight='bold')
            ax1.set_ylabel(self.moneda.get(), fontsize=12)
            ax1.grid(axis='y', alpha=0.3)
            
            # Gráfica 2: TCO Desglosado
            energia = metricas.get("costo_energia_anual_usd", 0) * 5
            mantenimiento = metricas.get("opex_mantenimiento", 0) * 5
            personal = metricas.get("opex_personal", 0) * 5
            
            categorias = ['CAPEX', 'Energía\n5 años', 'Mantenimiento\n5 años', 'Personal\n5 años']
            valores = [capex, energia, mantenimiento, personal]
            colores_barras = ['#00AEEF', '#f59e0b', '#ef4444', '#8b5cf6']
            
            ax2.bar(categorias, valores, color=colores_barras)
            ax2.set_title("TCO Desglosado (5 años)", fontsize=14, fontweight='bold')
            ax2.set_ylabel("USD", fontsize=12)
            ax2.grid(axis='y', alpha=0.3)
            
            plt.tight_layout()
            
            fecha = datetime.now().strftime("%Y%m%d_%H%M%S")
            ruta = RUTA_REPORTES / f"grafica_{fecha}.png"
            plt.savefig(ruta, dpi=150, bbox_inches='tight')
            plt.close()
            
            if os.name == "nt":
                os.startfile(str(ruta))
            else:
                os.system(f"open {ruta}")
            
            self._agregar_mensaje("sistema", f"✅ Gráfica generada: {ruta.name}", "exito")
        
        except Exception as e:
            self._agregar_mensaje("sistema", f"❌ Error: {e}", "error")
            traceback.print_exc()
    
    def _exportar_excel(self):
        """Exportar a Excel con desglose completo"""
        if not OPENPYXL_AVAILABLE:
            self._agregar_mensaje("sistema", "⚠️ Instala openpyxl: pip install openpyxl", "advertencia")
            return
        
        if not self.contexto.ultima_arquitectura:
            self._agregar_mensaje("sistema", "❌ Primero genera una arquitectura", "error")
            return
        
        try:
            arq = self.contexto.ultima_arquitectura.get("arquitectura", {})
            metricas = self.contexto.ultima_arquitectura.get("metricas", {})
            
            fecha = datetime.now().strftime("%Y%m%d_%H%M%S")
            ruta = RUTA_REPORTES / f"datacenter_{fecha}.xlsx"
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Arquitectura v2.1"
            
            # Título
            ws["A1"] = "Reporte de Arquitectura Data Center v2.1"
            ws["A1"].font = Font(bold=True, size=16)
            ws.merge_cells("A1:B1")
            
            # Datos
            row = 3
            datos = [
                ("Fecha", datetime.now().strftime("%d/%m/%Y %H:%M")),
                ("", ""),
                ("COMPONENTES", ""),
                ("CPU", f"{arq.get('cpu_model', '?')} ({arq.get('cores', 0)} cores)"),
                ("Motherboard", arq.get('mb_id', '?')),
                ("RAM", f"{arq.get('ram_gb', 0)} GB"),
                ("Storage", f"{arq.get('storage_type', '?')} - {arq.get('iops', 0):,} IOPS"),
                ("GPU", arq.get('gpu_model', 'Sin GPU')),
                ("Red", f"{arq.get('nic_model', '?')} ({arq.get('nic_gbps', 0)} Gbps)"),
                ("Enfriamiento", arq.get('cool_type', '?')),
                ("PSU", f"{arq.get('psu_watts', 0)} W"),
                ("UPS", f"{arq.get('ups_auto_min', 0)} min"),
                ("Chasis", f"{arq.get('rack_u', 0)} U"),
                ("", ""),
                ("ECONOMÍA", ""),
                ("CAPEX", f"${metricas.get('capex', 0):,.0f}"),
                ("OPEX Anual", f"${metricas.get('opex_anual', 0):,.0f}"),
                ("TCO 5 años", f"${metricas.get('tco_5anos', 0):,.0f}"),
                ("ROI", f"{metricas.get('roi_anios', 0):.1f} años"),
                ("", ""),
                ("EFICIENCIA", ""),
                ("Consumo", f"{metricas.get('consumo_por_nodo_w', 0):.0f} W"),
                ("PUE", f"{metricas.get('pue', 1.5):.2f}"),
                ("Consumo Anual", f"{metricas.get('consumo_anual_kwh', 0):,.0f} kWh"),
            ]
            
            for label, valor in datos:
                ws[f"A{row}"] = label
                ws[f"B{row}"] = valor
                if label in ["COMPONENTES", "ECONOMÍA", "EFICIENCIA"]:
                    ws[f"A{row}"].font = Font(bold=True, size=12)
                row += 1
            
            ws.column_dimensions["A"].width = 25
            ws.column_dimensions["B"].width = 50
            
            wb.save(ruta)
            
            if os.name == "nt":
                os.startfile(str(ruta))
            else:
                os.system(f"open {ruta}")
            
            self._agregar_mensaje("sistema", f"✅ Excel generado: {ruta.name}", "exito")
        
        except Exception as e:
            self._agregar_mensaje("sistema", f"❌ Error: {e}", "error")
            traceback.print_exc()
    
    def _guardar_json(self):
        """Guardar arquitectura en JSON"""
        if not self.contexto.ultima_arquitectura:
            self._agregar_mensaje("sistema", "❌ Primero genera una arquitectura", "error")
            return
        
        try:
            fecha = datetime.now().strftime("%Y%m%d_%H%M%S")
            ruta = RUTA_HISTORIAL / f"arquitectura_{fecha}.json"
            
            with open(ruta, "w", encoding="utf-8") as f:
                json.dump(self.contexto.ultima_arquitectura, f, indent=2, ensure_ascii=False, default=str)
            
            self._agregar_mensaje("sistema", f"✅ Guardado: {ruta.name}", "exito")
        
        except Exception as e:
            self._agregar_mensaje("sistema", f"❌ Error: {e}", "error")
    
    def _nuevo_diseno(self):
        """Reiniciar diseño"""
        self.contexto.limpiar()
        
        # Limpiar chat
        self.chat_box.configure(state="normal")
        self.chat_box.delete("1.0", "end")
        self.chat_box.configure(state="disabled")
        
        # Limpiar panel de componentes
        for widget in self.componentes_frame.winfo_children():
            widget.destroy()
        
        self.label_vacio.pack(pady=50)
        
        self._mensaje_bienvenida()
        self._agregar_mensaje("sistema", "🆕 Nuevo diseño iniciado. Contexto reiniciado.", "exito")
    
    def _refrescar_moneda(self, _=None):
        """Refrescar visualización al cambiar moneda"""
        if self.contexto.ultima_arquitectura:
            arq = self.contexto.ultima_arquitectura.get("arquitectura", {})
            metricas = self.contexto.ultima_arquitectura.get("metricas", {})
            self._renderizar_tarjetas(arq, metricas)
            self._agregar_mensaje("sistema", f"💱 Moneda cambiada a {self.moneda.get()}", "info")

# ============================================================
# PUNTO DE ENTRADA
# ============================================================

if __name__ == "__main__":
    app = SistemaExpertoDataCenter()
    app.mainloop()